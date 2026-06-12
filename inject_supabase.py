"""
inject_supabase.py
Crea la tabla en Supabase e inyecta todos los datos del Excel.
Uso: ejecutar inject_supabase.bat  (instala dependencias automáticamente)
"""

import sys, os, re
from datetime import datetime

# ── Credenciales ────────────────────────────────────────────────────────────
PG_URL   = "postgres://postgres.leputhlygxdxheayzpit:mTQ4zqojkd9BDdlR@aws-1-us-east-1.pooler.supabase.com:5432/postgres?sslmode=require"
EXCEL    = os.path.join(os.path.dirname(__file__), "Citas Contact Center 202620 10-06-2026.xlsx")
SHEET    = "Hoja1"
TABLE    = "citas_callcenter"
CHUNK    = 200   # filas por batch INSERT

# ── SQL: crear tabla ────────────────────────────────────────────────────────
CREATE_SQL = f"""
CREATE TABLE IF NOT EXISTS {TABLE} (
  id                    BIGSERIAL PRIMARY KEY,
  rut                   TEXT,
  asunto                TEXT,
  equipo                TEXT,
  cont_llamadas         INTEGER,
  fecha_creacion        DATE,
  fecha_compromiso      DATE,
  primer_sub_origen     TEXT,
  telefono              TEXT,
  propietario           TEXT,
  campus                TEXT,
  regimen               TEXT,
  carrera               TEXT,
  consultor             TEXT,
  propietario_contacto  TEXT,
  ultimo_interes_cierre TEXT,
  seguimiento_cierre    TEXT,
  interes               TEXT,
  interes_cierre        TEXT,
  prioridad             TEXT,
  semana                TEXT,
  mes                   TEXT,
  banner                TEXT,
  created_at            TIMESTAMPTZ DEFAULT NOW()
);
CREATE INDEX IF NOT EXISTS idx_{TABLE}_consultor ON {TABLE}(consultor);
CREATE INDEX IF NOT EXISTS idx_{TABLE}_campus    ON {TABLE}(campus);
CREATE INDEX IF NOT EXISTS idx_{TABLE}_carrera   ON {TABLE}(carrera);
CREATE INDEX IF NOT EXISTS idx_{TABLE}_estado    ON {TABLE}(ultimo_interes_cierre);
CREATE INDEX IF NOT EXISTS idx_{TABLE}_equipo    ON {TABLE}(equipo);
CREATE INDEX IF NOT EXISTS idx_{TABLE}_mes       ON {TABLE}(mes);
CREATE INDEX IF NOT EXISTS idx_{TABLE}_semana    ON {TABLE}(semana);
CREATE INDEX IF NOT EXISTS idx_{TABLE}_fecha     ON {TABLE}(fecha_creacion);
"""

# ── Mapeo columnas Excel → columnas Supabase ────────────────────────────────
def norm(s):
    import unicodedata
    s = unicodedata.normalize("NFD", str(s or "")).encode("ascii","ignore").decode().lower().strip()
    return re.sub(r"\s+", " ", s)

COL_MAP = {
    "rut":                                                  "rut",
    "asunto":                                               "asunto",
    "marketing5":                                           "equipo",
    "cont. llamadas (referente a) (contacto)":              "cont_llamadas",
    "fecha de cracion":                                     "fecha_creacion",
    "fecha de creacion":                                    "fecha_creacion",
    "fecha compromiso":                                     "fecha_compromiso",
    "primer sub origen (referente a) (contacto)":           "primer_sub_origen",
    "telefono movil (referente a) (contacto)":              "telefono",
    "propietario":                                          "propietario",
    "campus":                                               "campus",
    "regimen interes 1":                                    "regimen",
    "carrera de interes":                                   "carrera",
    "llamada a":                                            "consultor",
    "propietario (referente a) (contacto)":                 "propietario_contacto",
    "ultimo interes cierre (referente a) (contacto)":       "ultimo_interes_cierre",
    "seguimiento cierre":                                   "seguimiento_cierre",
    "interes":                                              "interes",
    "interes cierre":                                       "interes_cierre",
    "prioridad":                                            "prioridad",
    "semana":                                               "semana",
    "mes":                                                  "mes",
    "banner":                                               "banner",
}

def map_col(raw_header):
    n = norm(raw_header)
    return COL_MAP.get(n)

def clean_val(v, col):
    if v is None or v == "" or (isinstance(v, float) and str(v) == "nan"):
        return None
    if col in ("cont_llamadas",):
        try: return int(float(v))
        except: return None
    if col in ("fecha_creacion", "fecha_compromiso"):
        if isinstance(v, datetime):
            return v.date().isoformat()
        try:
            d = datetime.strptime(str(v)[:10], "%Y-%m-%d")
            return d.date().isoformat()
        except:
            return None
    return str(v).strip() or None

# ── Main ────────────────────────────────────────────────────────────────────
def main():
    try:
        import psycopg2
    except ImportError:
        print("Instalando psycopg2..."); os.system(f"{sys.executable} -m pip install psycopg2-binary -q"); import psycopg2

    try:
        import openpyxl
    except ImportError:
        print("Instalando openpyxl..."); os.system(f"{sys.executable} -m pip install openpyxl -q"); import openpyxl

    # Buscar Excel automáticamente si no está en el directorio del script
    excel_path = EXCEL
    if not os.path.exists(excel_path):
        # Buscar en escritorio y carpetas comunes
        search_dirs = [
            os.path.dirname(__file__),
            os.path.expanduser("~/Desktop"),
            os.path.expanduser("~/Downloads"),
            r"C:\Users\dcantillana\AppData\Roaming\Claude\local-agent-mode-sessions",
        ]
        for d in search_dirs:
            for root, _, files in os.walk(d):
                for f in files:
                    if "Citas Contact Center" in f and f.endswith(".xlsx"):
                        excel_path = os.path.join(root, f)
                        break
                if os.path.exists(excel_path): break
            if os.path.exists(excel_path): break

    if not os.path.exists(excel_path):
        print(f"\nERROR: No se encontró el Excel.")
        print("Pon 'Citas Contact Center 202620 10-06-2026.xlsx' en la misma carpeta que este script.")
        input("\nPresiona Enter para cerrar...")
        sys.exit(1)

    print(f"\n=== Inyector UDLA Call Center → Supabase ===")
    print(f"Excel: {excel_path}")
    print(f"Tabla: {TABLE}")

    # Conectar a PostgreSQL
    print("\n[1/4] Conectando a Supabase PostgreSQL...")
    conn = psycopg2.connect(PG_URL)
    conn.autocommit = True
    cur  = conn.cursor()
    print("  ✓ Conexión establecida")

    # Crear tabla e índices
    print("\n[2/4] Creando tabla e índices...")
    cur.execute(CREATE_SQL)
    print("  ✓ Tabla lista")

    # Leer Excel
    print("\n[3/4] Leyendo Excel...")
    import openpyxl
    wb   = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    ws   = wb[SHEET] if SHEET in wb.sheetnames else wb.active
    rows = list(ws.iter_rows(values_only=True))
    headers   = [str(h or "") for h in rows[0]]
    data_rows = rows[1:]
    print(f"  ✓ {len(data_rows)} filas · {len(headers)} columnas")

    # Mapear columnas
    col_idx = {}  # supabase_col → índice en headers
    for i, h in enumerate(headers):
        sb_col = map_col(h)
        if sb_col:
            col_idx[sb_col] = i

    print(f"  Columnas mapeadas: {', '.join(col_idx.keys())}")

    # Borrar datos previos (opcional — comenta si quieres append)
    print("\n  Limpiando datos previos...")
    cur.execute(f"TRUNCATE TABLE {TABLE} RESTART IDENTITY;")

    # Insertar en batches
    print("\n[4/4] Insertando datos...")
    sb_cols    = list(col_idx.keys())
    placeholders = ", ".join(["%s"] * len(sb_cols))
    col_list   = ", ".join(sb_cols)
    insert_sql = f"INSERT INTO {TABLE} ({col_list}) VALUES ({placeholders})"

    inserted = 0
    batch    = []
    for row in data_rows:
        values = []
        for col in sb_cols:
            raw = row[col_idx[col]] if col_idx[col] < len(row) else None
            values.append(clean_val(raw, col))
        batch.append(values)
        if len(batch) >= CHUNK:
            cur.executemany(insert_sql, batch)
            inserted += len(batch)
            batch = []
            print(f"  → {inserted} filas insertadas...", end="\r")

    if batch:
        cur.executemany(insert_sql, batch)
        inserted += len(batch)

    cur.close(); conn.close()

    print(f"\n\n=== ¡LISTO! ===")
    print(f"✓ {inserted} registros cargados en '{TABLE}'")
    print(f"✓ Tabla: https://supabase.com/dashboard/project/leputhlygxdxheayzpit/editor")
    input("\nPresiona Enter para cerrar...")

if __name__ == "__main__":
    main()
